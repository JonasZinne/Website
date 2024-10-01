import os
import pandas as pd
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

TEAM_SELECTOR = '.match-table-item__title'
DATE_SELECTOR = '.match-table-item__date'

def scroll_to_bottom(page):
    previous_height = page.evaluate("document.body.scrollHeight")
    while True:
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(500)
        new_height = page.evaluate("document.body.scrollHeight")
        if new_height == previous_height:
            break
        previous_height = new_height

def parse_html_content(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    team_divs = soup.select(TEAM_SELECTOR)
    date_divs = soup.select(DATE_SELECTOR)

    teams = [team.text for team in team_divs]
    dates = [date.text for date in date_divs]

    structured_data = []
    for i in range(len(dates)):
        team1 = teams[2 * i]
        team2 = teams[2 * i + 1]
        
        date, time = dates[i].split(' ')[0], dates[i].split(' ')[1]
        structured_data.append({'Date': date, 'Time': time, 'Team 1': team1, 'Team 2': team2})

    return structured_data

def style_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    center_alignment = Alignment(horizontal='center', vertical='center')

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4a90e2", end_color="4a90e2", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Border
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Columns
    ws.column_dimensions['A'].width = 15  # Datum
    ws.column_dimensions['B'].width = 10  # Uhrzeit
    ws.column_dimensions['C'].width = 25  # Team 1
    ws.column_dimensions['D'].width = 25  # Team 2

    # Datacells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border

    wb.save(file_path)

def main(url):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        with browser.new_page() as page:
            page.goto(url)
            scroll_to_bottom(page)
            content = page.content()

    structured_data = parse_html_content(content)
    df = pd.DataFrame(structured_data)

    storage_dir = 'storage'
    os.makedirs(storage_dir, exist_ok=True)
    excel_file = os.path.join(storage_dir, 'Matches.xlsx')
    df.to_excel(excel_file, index=False)

    style_excel(excel_file)

    return excel_file